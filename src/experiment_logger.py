import os
import json
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


EXCEL_FILE = "FER2013_Experiment_Tracking.xlsx"


EXPERIMENT_ID = "EXP-001"



# ============================================================
# CREATE EXCEL TEMPLATE
# ============================================================


def create_excel():

    if os.path.exists(EXCEL_FILE):
        return


    wb = Workbook()


    # -------------------------------
    # Summary Sheet
    # -------------------------------

    ws = wb.active
    ws.title = "Experiment Summary"


    ws.append([
        "Experiment ID",
        "Model",
        "Dataset",
        "Epochs",
        "Batch Size",
        "Optimizer",
        "Learning Rate",
        "Weight Decay",
        "GPU",
        "Best Epoch",
        "Best Val Accuracy",
        "Test Accuracy",
        "Precision",
        "Recall",
        "F1 Score"
    ])



    # -------------------------------
    # Class Performance
    # -------------------------------

    ws2 = wb.create_sheet(
        "Class Performance"
    )


    ws2.append([
        "Experiment ID",
        "Class",
        "Precision",
        "Recall",
        "F1 Score",
        "Support"
    ])



    # -------------------------------
    # Confusion Matrix
    # -------------------------------


    wb.create_sheet(
        "Confusion Matrix"
    )



    wb.save(EXCEL_FILE)





# ============================================================
# LOAD JSON RESULTS
# ============================================================


def load_results():


    with open(
        "outputs/results/test_results.json",
        "r"
    ) as f:

        test_results = json.load(f)



    with open(
        "outputs/results/training_history.json",
        "r"
    ) as f:

        history = json.load(f)



    return test_results, history






# ============================================================
# ADD EXPERIMENT
# ============================================================



def add_experiment():


    wb = load_workbook(
        EXCEL_FILE
    )


    test_results, history = load_results()



    # -------------------------------
    # Find Best Epoch
    # -------------------------------


    best_epoch = (
        history["val_accuracy"].index(
            max(history["val_accuracy"])
        )
        + 1
    )


    best_val_acc = max(
        history["val_accuracy"]
    )



    # -------------------------------
    # Summary
    # -------------------------------


    ws = wb[
        "Experiment Summary"
    ]



    ws.append([

        EXPERIMENT_ID,

        "EfficientNetV2-S",

        "FER2013",

        len(history["train_loss"]),

        32,

        "AdamW",

        0.0001,

        0.0001,

        "RTX 4090",

        best_epoch,

        best_val_acc,

        test_results["accuracy"],

        test_results["precision"],

        test_results["recall"],

        test_results["f1"]

    ])




    # -------------------------------
    # Class Results
    # -------------------------------


    ws = wb[
        "Class Performance"
    ]



    for item in test_results["classes"]:


        ws.append([

            EXPERIMENT_ID,

            item["class"],

            item["precision"],

            item["recall"],

            item["f1"],

            item["support"]

        ])




    # -------------------------------
    # Confusion Matrix
    # -------------------------------


    ws = wb[
        "Confusion Matrix"
    ]


    labels = [
        "angry",
        "disgust",
        "fear",
        "happy",
        "neutral",
        "sad",
        "surprise"
    ]


    ws.append(
        [
            "",
            *labels
        ]
    )


    for label,row in zip(
        labels,
        test_results["confusion_matrix"]
    ):


        ws.append(
            [
                label,
                *row
            ]
        )




    # -------------------------------
    # Auto Column Width
    # -------------------------------


    for sheet in wb:


        for column in sheet.columns:


            max_length = max(

                len(
                    str(cell.value)
                )
                if cell.value
                else 0

                for cell in column
            )


            sheet.column_dimensions[
                get_column_letter(
                    column[0].column
                )
            ].width = max_length + 3



    wb.save(
        EXCEL_FILE
    )



# ============================================================
# MAIN
# ============================================================


if __name__ == "__main__":


    create_excel()

    add_experiment()


    print(
        "Experiment tracking updated"
    )

    print(
        EXCEL_FILE
    )