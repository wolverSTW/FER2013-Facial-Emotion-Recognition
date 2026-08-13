"""
FER2013 Facial Emotion Recognition
==================================

Final evaluation pipeline.

Features:
- Evaluate best trained model
- Official FER2013 test set
- Accuracy
- Precision
- Recall
- F1-score
- Classification report
- Confusion matrix
- Save JSON result
- Save TXT report
- Update Excel experiment tracker

"""


# ============================================================
# 1. IMPORT LIBRARIES
# ============================================================


import json

from pathlib import Path


import torch

from torch import nn


from sklearn.metrics import (
    accuracy_score,
    classification_report,
    confusion_matrix,
    precision_recall_fscore_support,
)


from openpyxl import (
    Workbook,
    load_workbook,
)


from openpyxl.utils import (
    get_column_letter,
)



# ============================================================
# 2. PROJECT IMPORTS
# ============================================================


from src.config import (
    DEVICE,
    MODEL_DIR,
    CLASS_NAMES,
)


from src.dataset import (
    load_datasets,
    validate_class_mapping,
    create_dataloaders,
)


from src.model import create_model




# ============================================================
# 3. EVALUATION FUNCTION
# ============================================================


@torch.no_grad()
def evaluate_model(
    model,
    test_loader,
    device,
):

    model.eval()


    criterion = nn.CrossEntropyLoss()


    running_loss = 0.0

    total_samples = 0


    predictions_all = []

    labels_all = []



    for images, labels in test_loader:


        images = images.to(
            device
        )


        labels = labels.to(
            device
        )


        outputs = model(
            images
        )


        loss = criterion(
            outputs,
            labels
        )


        running_loss += (
            loss.item()
            *
            images.size(0)
        )


        total_samples += (
            labels.size(0)
        )



        predictions = outputs.argmax(
            dim=1
        )


        predictions_all.extend(
            predictions.cpu().numpy()
        )


        labels_all.extend(
            labels.cpu().numpy()
        )



    test_loss = (
        running_loss
        /
        total_samples
    )



    accuracy = accuracy_score(
        labels_all,
        predictions_all
    )


    precision, recall, f1, _ = (
        precision_recall_fscore_support(
            labels_all,
            predictions_all,
            average="weighted",
            zero_division=0
        )
    )


    report = classification_report(
        labels_all,
        predictions_all,
        target_names=CLASS_NAMES,
        output_dict=True,
        zero_division=0
    )


    cm = confusion_matrix(
        labels_all,
        predictions_all
    )



    return {


        "test_loss":
            float(test_loss),


        "test_accuracy":
            float(accuracy),


        "weighted_precision":
            float(precision),


        "weighted_recall":
            float(recall),


        "weighted_f1":
            float(f1),


        "classification_report":
            report,


        "confusion_matrix":
            cm.tolist()

    }


# ============================================================
# 4. SAVE JSON RESULTS
# ============================================================


def save_json_results(results):

    results_dir = (
        MODEL_DIR.parent
        /
        "results"
    )


    results_dir.mkdir(
        parents=True,
        exist_ok=True
    )


    path = (
        results_dir
        /
        "evaluation_results.json"
    )


    with open(
        path,
        "w",
        encoding="utf-8"
    ) as file:


        json.dump(
            results,
            file,
            indent=4
        )


    print(
        "JSON saved:"
    )

    print(
        path
    )





# ============================================================
# 5. SAVE TEXT RESULTS
# ============================================================


def save_text_results(results):


    results_dir = (
        MODEL_DIR.parent
        /
        "results"
    )


    path = (
        results_dir
        /
        "test_results.txt"
    )


    report = results[
        "classification_report"
    ]



    with open(
        path,
        "w",
        encoding="utf-8"
    ) as file:


        file.write(
            "FER2013 FINAL TEST RESULTS\n"
        )

        file.write(
            "=" * 60
            +
            "\n\n"
        )


        file.write(
            f"Test Loss       : "
            f"{results['test_loss']:.4f}\n"
        )


        file.write(
            f"Test Accuracy   : "
            f"{results['test_accuracy']:.4f}\n"
        )


        file.write(
            f"Precision       : "
            f"{results['weighted_precision']:.4f}\n"
        )


        file.write(
            f"Recall          : "
            f"{results['weighted_recall']:.4f}\n"
        )


        file.write(
            f"F1 Score        : "
            f"{results['weighted_f1']:.4f}\n\n"
        )


        file.write(
            "CLASS PERFORMANCE\n"
        )

        file.write(
            "-" * 60
            +
            "\n"
        )


        for cls in CLASS_NAMES:


            data = report[cls]


            file.write(

                f"{cls:<12}"
                f"{data['precision']:.4f}   "
                f"{data['recall']:.4f}   "
                f"{data['f1-score']:.4f}   "
                f"{int(data['support'])}\n"

            )



    print(
        "Text result saved:"
    )

    print(
        path
    )





# ============================================================
# 6. SAVE CONFUSION MATRIX
# ============================================================


def save_confusion_matrix(results):


    path = (
        MODEL_DIR.parent
        /
        "results"
        /
        "confusion_matrix.txt"
    )



    with open(
        path,
        "w",
        encoding="utf-8"
    ) as file:


        file.write(
            "Classes:\n"
        )


        file.write(
            str(CLASS_NAMES)
        )


        file.write(
            "\n\n"
        )


        file.write(
            "Confusion Matrix:\n"
        )


        for row in results[
            "confusion_matrix"
        ]:


            file.write(
                str(row)
                +
                "\n"
            )



    print(
        "Confusion matrix saved:"
    )

    print(
        path
    )





# ============================================================
# 7. UPDATE EXCEL TRACKER
# ============================================================


def update_excel_tracker(results):


    excel_path = (
        MODEL_DIR.parent.parent
        /
        "FER2013_Experiment_Tracking.xlsx"
    )


    if not excel_path.exists():


        wb = Workbook()


        ws = wb.active

        ws.title = (
            "Experiment Summary"
        )


        ws.append([

            "Experiment ID",
            "Model",
            "Dataset",
            "Epochs",
            "Optimizer",
            "Learning Rate",
            "Weight Decay",
            "GPU",
            "Best Epoch",
            "Validation Accuracy",
            "Test Accuracy",
            "Precision",
            "Recall",
            "F1 Score"

        ])



        ws2 = wb.create_sheet(
            "Class Performance"
        )


        ws2.append([

            "Experiment ID",
            "Class",
            "Precision",
            "Recall",
            "F1 Score",
            "Support"

        ])



        wb.save(
            excel_path
        )





    wb = load_workbook(
        excel_path
    )



    # Summary sheet

    ws = wb[
        "Experiment Summary"
    ]



    ws.append([

        "EXP-001",

        "EfficientNetV2-S",

        "FER2013",

        28,

        "AdamW",

        0.0001,

        0.0001,

        torch.cuda.get_device_name(0)
        if torch.cuda.is_available()
        else "CPU",

        23,

        0.7073,

        results["test_accuracy"],

        results["weighted_precision"],

        results["weighted_recall"],

        results["weighted_f1"]

    ])




    # Class sheet


    ws = wb[
        "Class Performance"
    ]


    report = results[
        "classification_report"
    ]



    for cls in CLASS_NAMES:


        ws.append([


            "EXP-001",

            cls,

            report[cls]["precision"],

            report[cls]["recall"],

            report[cls]["f1-score"],

            report[cls]["support"]

        ])





    # Auto column width


    for sheet in wb:


        for column in sheet.columns:


            max_length = max(

                len(str(cell.value))
                if cell.value
                else 0

                for cell in column

            )


            sheet.column_dimensions[
                get_column_letter(
                    column[0].column
                )
            ].width = (
                max_length + 3
            )



    wb.save(
        excel_path
    )


    print(
        "Excel tracker updated:"
    )

    print(
        excel_path
    )





# ============================================================
# 8. MAIN
# ============================================================


def main():


    print()

    print("=" * 70)

    print(
        "FER2013 FINAL MODEL EVALUATION"
    )

    print("=" * 70)



    print()

    print(
        f"Device: {DEVICE}"
    )


    if torch.cuda.is_available():

        print(
            f"GPU: "
            f"{torch.cuda.get_device_name(0)}"
        )



    # Load dataset


    print()

    print(
        "Loading datasets..."
    )


    train_dataset, test_dataset = (
        load_datasets()
    )



    validate_class_mapping(
        train_dataset,
        test_dataset
    )



    (
        _,
        _,
        test_loader
    ) = create_dataloaders(

        train_dataset,

        test_dataset,

        test_dataset

    )



    # Create model


    print()

    print(
        "Loading best model..."
    )


    model = create_model()



    model_path = (
        MODEL_DIR
        /
        "best_model.pth"
    )



    model.load_state_dict(

        torch.load(
            model_path,
            map_location=DEVICE
        )

    )



    model = model.to(
        DEVICE
    )



    # Evaluation


    print()

    print(
        "Evaluating..."
    )


    results = evaluate_model(

        model,

        test_loader,

        DEVICE

    )



    print()

    print("=" * 70)

    print(
        "FINAL TEST RESULTS"
    )

    print("=" * 70)



    print(

        f"Accuracy: "
        f"{results['test_accuracy']:.4f}"

    )


    print(

        f"F1 Score: "
        f"{results['weighted_f1']:.4f}"

    )



    # Save everything


    save_json_results(
        results
    )


    save_text_results(
        results
    )


    save_confusion_matrix(
        results
    )


    update_excel_tracker(
        results
    )



    print()

    print(
        "Evaluation completed."
    )





# ============================================================
# RUN
# ============================================================


if __name__ == "__main__":

    main()