"""
FER2013 Experiment Tracking
---------------------------

Creates an Excel experiment-tracking workbook and records
the first baseline experiment (EXP-001).

Experiment:
    EXP-001 - EfficientNetV2-S Baseline

The workbook is designed so that future experiments can be
added as new rows without changing the existing baseline.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 1. PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_DIR = PROJECT_ROOT / "experiment_logs"

EXCEL_PATH = (
    PROJECT_ROOT / "FER2013_Experiment_Tracking.xlsx"
)


# ============================================================
# 2. BASELINE EXPERIMENT INFORMATION
# ============================================================

EXPERIMENT_ID = "EXP-001"

experiment_summary = {
    "Experiment ID": EXPERIMENT_ID,
    "Experiment Type": "Baseline",
    "Dataset": "FER2013",
    "Model": "EfficientNetV2-S",
    "Input Size": "48x48",
    "Classes": 7,
    "Train Split": "90%",
    "Validation Split": "10%",
    "Test Set": "Official FER2013 Test",
    "Batch Size": 32,
    "Epochs Configured": 30,
    "Epochs Completed": 28,
    "Best Epoch": 23,
    "Learning Rate": 0.0001,
    "Weight Decay": 0.0001,
    "Optimizer": "AdamW",
    "Scheduler": "CosineAnnealingLR",
    "Loss Function": "CrossEntropyLoss",
    "Early Stopping Patience": 5,
    "Best Validation Accuracy": 0.7073,
    "Test Loss": 1.8210,
    "Test Accuracy": 0.7099,
    "Weighted Precision": 0.7102,
    "Weighted Recall": 0.7099,
    "Weighted F1": 0.7090,
    "Device": "NVIDIA GeForce RTX 4090",
    "CUDA": "12.8",
    "Notes": "Baseline model",
}


# ============================================================
# 3. CLASSIFICATION RESULTS
# ============================================================

class_results = [
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "angry",
        "Precision": 0.6663,
        "Recall": 0.6086,
        "F1-score": 0.6361,
        "Support": 958,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "disgust",
        "Precision": 0.7449,
        "Recall": 0.6577,
        "F1-score": 0.6986,
        "Support": 111,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "fear",
        "Precision": 0.6169,
        "Recall": 0.5332,
        "F1-score": 0.5720,
        "Support": 1024,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "happy",
        "Precision": 0.8841,
        "Recall": 0.8856,
        "F1-score": 0.8848,
        "Support": 1774,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "neutral",
        "Precision": 0.6324,
        "Recall": 0.6991,
        "F1-score": 0.6641,
        "Support": 1233,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "sad",
        "Precision": 0.5784,
        "Recall": 0.6151,
        "F1-score": 0.5962,
        "Support": 1247,
    },
    {
        "Experiment ID": EXPERIMENT_ID,
        "Class": "surprise",
        "Precision": 0.8126,
        "Recall": 0.8351,
        "F1-score": 0.8237,
        "Support": 831,
    },
]


# ============================================================
# 4. CONFUSION MATRIX
# ============================================================

classes = [
    "angry",
    "disgust",
    "fear",
    "happy",
    "neutral",
    "sad",
    "surprise",
]

confusion_matrix = [
    [583, 13, 91, 36, 84, 137, 14],
    [18, 73, 6, 2, 2, 9, 1],
    [90, 3, 546, 22, 95, 195, 73],
    [23, 3, 22, 1571, 92, 25, 38],
    [61, 2, 45, 74, 862, 173, 16],
    [84, 4, 120, 45, 209, 767, 18],
    [16, 0, 55, 27, 19, 20, 694],
]


# ============================================================
# 5. TRAINING INFORMATION
# ============================================================

training_results = {
    "Experiment ID": EXPERIMENT_ID,
    "Original Training Samples": 28709,
    "Training Samples": 25839,
    "Validation Samples": 2870,
    "Official Test Samples": 7178,
    "Training Batches": 808,
    "Validation Batches": 90,
    "Test Batches": 225,
    "Total Parameters": 20186455,
    "Trainable Parameters": 20186455,
    "Best Epoch": 23,
    "Best Validation Accuracy": 0.7073,
    "Epochs Completed": 28,
    "Early Stopping Patience": 5,
}


# ============================================================
# 6. STYLING
# ============================================================

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

header_font = Font(
    color="FFFFFF",
    bold=True,
)

title_font = Font(
    bold=True,
    size=14,
)

bold_font = Font(
    bold=True,
)

thin_side = Side(
    style="thin",
    color="B7B7B7",
)

border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side,
)


# ============================================================
# 7. HELPER FUNCTIONS
# ============================================================

def style_header(ws):
    """Apply formatting to the first row."""

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.border = border


def auto_width(ws):
    """Automatically adjust worksheet column widths."""

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                value_length = len(
                    str(cell.value)
                )

                max_length = max(
                    max_length,
                    value_length,
                )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            40,
        )


def format_sheet(ws):

    for row in ws.iter_rows():

        for cell in row:

            cell.border = border

            cell.alignment = Alignment(
                vertical="center",
            )

    auto_width(ws)


# ============================================================
# 8. CREATE WORKBOOK
# ============================================================

def create_workbook():

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # Remove default sheet.
    default_sheet = workbook.active

    workbook.remove(default_sheet)


    # ========================================================
    # SHEET 1: EXPERIMENT SUMMARY
    # ========================================================

    ws = workbook.create_sheet(
        "Experiment Summary"
    )

    headers = list(
        experiment_summary.keys()
    )

    ws.append(headers)

    ws.append(
        list(
            experiment_summary.values()
        )
    )

    style_header(ws)

    format_sheet(ws)

    # Percentage formatting.
    percentage_fields = {
        "Best Validation Accuracy",
        "Weighted Precision",
        "Weighted Recall",
        "Weighted F1",
    }

    for column_index, header in enumerate(
        headers,
        start=1,
    ):

        if header in percentage_fields:

            ws.cell(
                row=2,
                column=column_index,
            ).number_format = "0.00%"


    # ========================================================
    # SHEET 2: CLASS RESULTS
    # ========================================================

    ws = workbook.create_sheet(
        "Class Results"
    )

    headers = [
        "Experiment ID",
        "Class",
        "Precision",
        "Recall",
        "F1-score",
        "Support",
    ]

    ws.append(headers)

    for result in class_results:

        ws.append(
            [
                result["Experiment ID"],
                result["Class"],
                result["Precision"],
                result["Recall"],
                result["F1-score"],
                result["Support"],
            ]
        )

    style_header(ws)

    format_sheet(ws)

    for row in range(
        2,
        ws.max_row + 1,
    ):

        ws.cell(
            row=row,
            column=3,
        ).number_format = "0.00%"

        ws.cell(
            row=row,
            column=4,
        ).number_format = "0.00%"

        ws.cell(
            row=row,
            column=5,
        ).number_format = "0.00%"


    # ========================================================
    # SHEET 3: CONFUSION MATRIX
    # ========================================================

    ws = workbook.create_sheet(
        "Confusion Matrix"
    )

    ws.append(
        [
            "Actual \\ Predicted"
        ]
        + classes
    )

    for actual_class, values in zip(
        classes,
        confusion_matrix,
    ):

        ws.append(
            [
                actual_class
            ]
            + values
        )

    style_header(ws)

    format_sheet(ws)

    # Center confusion matrix.
    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


    # ========================================================
    # SHEET 4: Training Information
    # ========================================================

    ws = workbook.create_sheet(
        "Training Information"
    )

    ws.append(
        [
            "Experiment ID",
            "Metric",
            "Value",
        ]
    )

    for metric, value in training_results.items():

        if metric == "Experiment ID":
            continue

        ws.append(
            [
                EXPERIMENT_ID,
                metric,
                value,
            ]
        )

    style_header(ws)

    format_sheet(ws)

    # Format accuracy.
    for row in range(
        2,
        ws.max_row + 1,
    ):

        metric = ws.cell(
            row=row,
            column=2,
        ).value

        if metric == "Best Validation Accuracy":

            ws.cell(
                row=row,
                column=3,
            ).number_format = "0.00%"


    # ========================================================
    # SHEET 5: README
    # ========================================================

    ws = workbook.create_sheet(
        "README",
        0,
    )

    ws["A1"] = (
        "FER2013 Experiment Tracking"
    )

    ws["A1"].font = title_font

    ws["A3"] = "Purpose"

    ws["A3"].font = bold_font

    ws["A4"] = (
        "This workbook tracks FER2013 experiments "
        "and provides EXP-001 as the baseline reference."
    )

    ws["A6"] = "Baseline Experiment"

    ws["A6"].font = bold_font

    ws["A7"] = (
        "EXP-001 - EfficientNetV2-S Baseline"
    )

    ws["A9"] = "Important"

    ws["A9"].font = bold_font

    ws["A10"] = (
        "Future experiments should be added as "
        "EXP-002, EXP-003, EXP-004, etc."
    )

    ws["A12"] = "Baseline Test Accuracy"

    ws["A12"].font = bold_font

    ws["B12"] = 0.7099

    ws["B12"].number_format = "0.00%"

    ws["A13"] = "Baseline Weighted F1"

    ws["A13"].font = bold_font

    ws["B13"] = 0.7090

    ws["B13"].number_format = "0.00%"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 65


    # ========================================================
    # SAVE
    # ========================================================

    workbook.save(
        EXCEL_PATH
    )

    print()
    print("=" * 70)
    print("EXPERIMENT TRACKING CREATED")
    print("=" * 70)
    print()
    print(
        f"Excel file: {EXCEL_PATH}"
    )
    print()
    print(
        "Baseline experiment: EXP-001"
    )
    print(
        "Model: EfficientNetV2-S"
    )
    print(
        "Validation Accuracy: 70.73%"
    )
    print(
        "Test Accuracy: 70.99%"
    )
    print(
        "Weighted F1: 70.90%"
    )
    print()
    print(
        "Sheets created:"
    )
    print(
        "  1. README"
    )
    print(
        "  2. Experiment Summary"
    )
    print(
        "  3. Class Results"
    )
    print(
        "  4. Confusion Matrix"
    )
    print(
        "  5. Training Information"
    )
    print()


# ============================================================
# 9. MAIN
# ============================================================

if __name__ == "__main__":

    create_workbook()
